from tkinter import ttk
from tkinter import *
from tkinter import messagebox

from datetime import date


bill = Tk()
bill.title("omsakthi adagu kadai")
bill.geometry("1000x1000")
bill.configure(bg="#535c68")
bill.state("zoomed")

head_frame=Frame(bill,bg="#535c68")
head_frame.pack(side=TOP)

lblfont="Calibri",16
todate="date"

bill1=StringVar()
name1=StringVar()
date1=StringVar()
coname1=StringVar()
street1=StringVar()
address1=StringVar()
amount1=StringVar()
item1=StringVar()
weight1=StringVar()
noitem1=StringVar()
relese1=StringVar()
intrel=StringVar()
phvar=StringVar()

#####
from openpyxl import load_workbook,worksheet,workbook,Workbook
import numpy as np
import os,sys
from docxtpl import DocxTemplate
from tkinter import *
from tkcalendar import *
from datetime import date,datetime
import mysql.connector as ms

def resource(relative_path):
     try:
          base_path=sys._MEIPASS
     except Exception:
           base_path=os.path.abspath(".")
     return os.path.join(base_path,relative_path)

try:
    wb=load_workbook(resource("src/new.xlsx"))
    ws=wb.active
    if ws["A1"].value=="loan_date":
        pass
    else:
        head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
        ws.append(head)
        wb.save(resource("src/new.xlsx"))
    
except:
    wb=Workbook()
    ws=wb.active
    head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
    ws.append(head)
    wb.save(resource("src/new.xlsx")) 



con=ms.connect(host="localhost",user="root",password="root")
cur=con.cursor()

try:
    cur.execute("create database oms")
    cur.execute("use oms")
except:
    cur.execute("use oms")
    try:
        cur.execute("create table omm (bill_no int primary key,loan_date varchar(50),name varchar(50),co_name varchar(50),street varchar(50),address varchar(50),int_amt int,weigth int,item varchar(50),no_item int,Phone_no varchar(50),releas varchar(50))")
    except:
        pass
con.commit()

def dbadd(lis):
    cur.execute('insert into omm (bill_no,loan_date,name,co_name,street,address,int_amt,weigth,item,no_item,Phone_no) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(lis[0],lis[1],lis[2],lis[3],lis[4],lis[5],lis[6],lis[7],lis[8],lis[9],lis[10]))
    con.commit()
    
def dbcheck(b):
    cur.execute('select max(bill_no) from omm')    
    b1=(cur.fetchone())[0]
    if b1>b:
        return b1
    else:
        return b
    



def save(a,lis):#for save not print or edit
    ws.append(a)
    wb.save(resource("src/new.xlsx"))
    dbadd(lis)
    return True

def loanprint(a,c,l):
    bill=a[1]
    date=a[0]
    name=a[2]
    coname=a[3]
    street=a[4]
    address=a[5]
    item=a[8]
    weight=a[7]
    amount=a[6]
    noitem=a[9]
    phnum=a[10]
    dblis=(bill,date,name,coname,address,street,amount,weight,item,noitem,phnum)
    if l=="loan":
        save(a,dblis)
    maxd=datechange(date)
    m=datetime.strptime((maxd),'%d-%m-%Y')
    max_date=(m.replace(year=m.year+1))
    max_date=max_date.strftime('%d-%m-%Y')
    newloan=DocxTemplate(resource('src/loan.docx')) 
    newloan.render({'loan_day':date,'bill_no':bill,'name':name,'coname':coname,'address':address,'street':street,'item':item,'weight':weight,'amount':amount,'noitem':noitem,'max_date':max_date,'Ph_no':phnum})
    loan_name=f"{bill} {name} {l}.docx"
    if l=="loan":
        loan_path=r'.\loan'
    elif l=="reprint":
        loan_path=r'.\reprint'
    n_path=resource(os.path.join(loan_path,loan_name))
    newloan.save(n_path)
    if c==True:
        filepath=loan_path
        os.startfile(n_path,'print')


def interestprint(paylist,c):
    doc=DocxTemplate(resource('src/omm.docx'))  
    name=paylist[0]
    loan_date=paylist[1]
    bill_no=paylist[2]
    int_amt=paylist[3]
    to_day=paylist[4]
    interest=paylist[5]
    total=paylist[6]

    id=int(paylist[7])+1
    ws[f"l{id}"]=to_day
    wb.save(resource("src/new.xlsx"))
    cur.execute('update omm set releas=%s where bill_no=%s',(to_day,bill_no))
    con.commit()

    doc.render({'name':name,'to_day':to_day,'loan_date':loan_date,'bill_no':bill_no,'interest':interest,'total':total,'int_amt':int_amt})
    r_name=f"{bill_no} {name}.docx"
    patth=r'.\relese'
    r_path=resource(os.path.join(patth,r_name))
    doc.save(r_path)
    if c==True:
        filepath=r_path
        os.startfile(filepath,'print')

def max_bill():
    s=list(ws.columns)[1]
    c=[]
    for i in range(1,ws.max_row):
        try:
            a=str(s[i].value)
            c.append(int(a))
        except:
            pass
    c.sort()
    m=dbcheck(c[-1])
    return (m+1)



def alter(id,uplist):
        
        id+=1
        name=ws[f"C{id}"]=uplist[2]
        co_name=ws[f"D{id}"]=uplist[3]
        loan_date=ws[f"A{id}"]=uplist[0]
        bill_no=ws[f"B{id}"]=uplist[1]
        address=ws[f"F{id}"]=uplist[5]
        weight=ws[f"H{id}"]=uplist[7]
        item=ws[f"I{id}"]=uplist[8]
        int_amt=ws[f"G{id}"]=uplist[6]
        street=ws[f"E{id}"]=uplist[4]
        no_item=ws[f"J{id}"]=uplist[9]
        Phone_no=ws[f"k{id}"]=uplist[10]
        releas=ws[f"l{id}"]=uplist[11]
        
        wb.save(resource("src/new.xlsx"))

        cur.execute('update omm set loan_date=%s,name=%s,co_name=%s,street=%s,address=%s,int_amt=%s,weigth=%s,item=%s,no_item=%s,Phone_no=%s,releas=%s where bill_no=%s',(loan_date,name,co_name,street,address,int_amt,weight,item,no_item,Phone_no,releas,bill_no))
        con.commit()
    


def srch(i,val):
    
        id=val
        i=str(i)
        if i=="கடன் தேதி":
            src='loan_date'
        elif i=="கடன் எண்":
            src='bill_no'
        elif i=="பெயர்":
            src='name'
        elif i=="த/க பெயர்":
            src='co_name'
        elif i=="ஊர்":
            src='address'
        elif i=="கடன் தொகை":
            src='amount'
        elif i=="பொருள்":
            src='items'
        elif i=="மீட்ட தேதி":
            src='relese_date'
        elif i=="எடை":
            src='weight'
        elif i=='Phone No':
            src="Phone No"
        else:
            src="none"
        

        
        op=[]
        c=False
        x=[]
        if src=='loan_date':
            id=datechange(id)
            try:
                s=list(ws.columns)[0]
                for i in range(1,(ws.max_row)):
                    if datechange(s[i].value)==(id):
                        op.append(i)
                    else:
                        raise ValueError

            except ValueError:
                cur.execute('select * from omm where loan_date=%s',(id,))
                x=cur.fetchall()
                
                print(x)
                c=True
                
        
    
        elif src=='bill_no':
            try:
                s=list(ws.columns)[1]
                for i in range(1,ws.max_row):
                    a=str(s[i].value)
                    if int(a)==int(id):
                        op.append(i)
                    else:
                        raise ValueError

            except ValueError:
                cur.execute('select * from omm where bill_no=%s',(id,))
                x=cur.fetchall()
                print(x)
             
                c=True


        elif src=='name':
            s=np.array(list(ws.columns)[2])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='co_name':
            s=list(ws.columns)[3]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='address':
            s=np.array(list(ws.columns)[5])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='amount':
            s=list(ws.columns)[6]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass

        elif src=='weight':
            s=list(ws.columns)[7]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass
        
        elif src=='items':
            s=list(ws.columns)[8]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='Phone No':
            s=list(ws.columns)[10]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='relese_date':
            id=datechange(id)
            s=list(ws.columns)[11]
            for i in range(1,ws.max_row):
                if datechange(s[i].value)==datechange(id):
                    op.append(i)

        if c==False:
            for a in op:
                c=list(ws.rows)[int(a)]
                v=[]
                for i in range(0,12):
                    # if c[i].value==None:
                    #     pass
                    v.append(c[i].value)
                v.append(a)
                x.append(v)
        print(x)
        return x

def delete(id):
    ws.delete_rows((int(id)+1))
    wb.save(resource("src/new.xlsx"))

    
def datechange(a):        
        t1="%d-%m-%y %H:%M:%S"
        t2="%d-%m-%y"
        t3="%d %m %y %H:%M:%S"
        t4="%d %m %y"
        t5="%d/%m/%y %H:%M:%S"
        t6="%d/%m/%y"
        t7="%d-%m-%Y %H:%M:%S"
        t8="%d-%m-%Y"
        t9="%d %m %Y %H:%M:%S"
        t10="%d %m %Y"
        t11="%d/%m/%Y %H:%M:%S"
        t12="%d/%m/%Y"
        t13="%Y-%m-%d %H:%M:%S"
        t14="%Y-%m-%d"
        t=[t1,t2,t3,t4,t5,t6,t7,t8,t9,t10,t11,t12,t13,t14]
        try:
            for i in t:
                try:
                    a=datetime.strptime(a,i)
                    
                except:
                    pass
            return a.strftime("%d-%m-%Y")
        except:
            return a

def interest(int_amt,loan_date,re_date):
    loan_date=datetime.strptime(loan_date,"%d-%m-%Y")
    re_date=datetime.strptime(re_date,"%d-%m-%Y")
    diff=re_date-loan_date
    intday=int(diff.days)
    if intday<15:
        intday=15
    intrest=(int(int_amt)*0.015*intday)/30
    total=int(intrest)+int(int_amt)
    
    return (int(intrest))  

#####

tday=(date.today()).strftime('%d-%m-%Y')
val=None

stlist=[bill1,name1,date1,coname1,street1,address1,amount1,item1,weight1,noitem1,relese1,intrel,phvar]

def stclear():
    global val
    for i in stlist:
        i.set("")
    val=None

def ltime(event):
    global datewindow,cal
    datewindow=Toplevel()
    datewindow.grab_set()
    datewindow.title("dd")
    datewindow.geometry("250x220+590+370")
    cal=Calendar(datewindow,selectmode="day",date_pattern="dd-mm-y")
    cal.place(x=0,y=0)
    bb=Button(datewindow,text="date",command=pic)
    bb.place(x=80,y=190)
    return cal.get_date()
def pic():
    date1.set(cal.get_date())
    datewindow.destroy()
    

def loan():
    global billing_frame
    try:
        search_frame.destroy()
        tree_frame.destroy()
        update_frame.destroy()
        tkint_frame.destroy()
    except:
        pass
    try:
        billing_frame.destroy()
    except:
        pass
    
    stclear()
    
    billing_frame=Frame(bill,bg="#535c68")
    billing_frame.pack(side=TOP,fill=X)

    title= Label(billing_frame, text="new loan", font=( "Calibri", 16, "bold"),bg="#535c68",height=0)
    title.grid(row=0, columnspan=2,padx=20,pady=20)

    lbldate=Label(billing_frame,text="தேதி",font=lblfont,bg="#535c68")
    lbldate.grid(row=1,column=1)
    txtdate=Entry(billing_frame,font=lblfont,width=20,textvariable=date1)
    txtdate.grid(row=1,column=2,pady=5)

    date1.set(tday)
    txtdate.bind("<1>",ltime)

    lblbill=Label(billing_frame,text="கடன் எண்",font=lblfont,bg="#535c68")
    lblbill.grid(row=2,column=1)
    txtbill=Entry(billing_frame,font=lblfont,width=20,textvariable=bill1)
    txtbill.grid(row=2,column=2,pady=5)
    try:
        bill1.set(max_bill())
    except:
        bill.set("000")




    lblname=Label(billing_frame,text="பெயர்",font=lblfont,bg="#535c68")
    lblname.grid(row=3,column=1)
    txtname=Entry(billing_frame,font=lblfont,width=20)
    txtname.grid(row=3,column=2,pady=5)


    lblconame=Label(billing_frame,text="த/க பெயர்",font=lblfont,bg="#535c68")
    lblconame.grid(row=4,column=1)
    txtconame=Entry(billing_frame,font=lblfont,width=20)
    txtconame.grid(row=4,column=2,pady=5)

    lblstreet=Label(billing_frame,text="தெரு",font=lblfont,bg="#535c68")
    lblstreet.grid(row=5,column=1)
    txtstreet=Entry(billing_frame,font=lblfont,width=20,textvariable=street1)
    txtstreet.grid(row=5,column=2,pady=5)

    lbladress=Label(billing_frame,text="ஊர்",font=lblfont,bg="#535c68")
    lbladress.grid(row=6,column=1)
    txtadress=Entry(billing_frame,font=lblfont,width=20)
    txtadress.grid(row=6,column=2,pady=5)

    lblamount=Label(billing_frame,text="கடன் தொகை",font=lblfont,bg="#535c68")
    lblamount.grid(row=7,column=1)
    txtamount=Entry(billing_frame,font=lblfont,width=20)
    txtamount.grid(row=7,column=2,pady=5)

    lblitem=Label(billing_frame,text="பொருள்",font=lblfont,bg="#535c68")
    lblitem.grid(row=8,column=1)
    txtitem=Entry(billing_frame,font=lblfont,width=20)
    txtitem.grid(row=8,column=2,pady=5)

    lblweight=Label(billing_frame,text="எடை",font=lblfont,bg="#535c68")
    lblweight.grid(row=9,column=1)
    txtweight=Entry(billing_frame,font=lblfont,width=20)
    txtweight.grid(row=9,column=2,pady=5)

    lblnoitem=Label(billing_frame,text="மொத்த பொருள்",font=lblfont,bg="#535c68")
    lblnoitem.grid(row=10,column=1)
    txtnoitem=Entry(billing_frame,font=lblfont,width=20)
    txtnoitem.grid(row=10,column=2,pady=5)

    lblph=Label(billing_frame,text="Phone No.",font=lblfont,bg="#535c68")
    lblph.grid(row=11,column=1)
    txtph=Entry(billing_frame,font=lblfont,width=20)
    txtph.grid(row=11,column=2,pady=5)

    def new_loan():
        bill=txtbill.get()
        date=txtdate.get()
        try:
            date=datechange(date)
        except:
            pass
        name=txtname.get()
        coname=txtconame.get()
        address=txtadress.get()
        item=txtitem.get()
        weight=txtweight.get()
        amount=txtamount.get()
        noitem=txtnoitem.get()
        street=txtstreet.get()
        phno=txtph.get()
        def n_loan():
            dlist=[date,bill,name,coname,address,amount,weight,item,noitem]
            plist=[date,bill,name,coname,street,address,amount,weight,item,noitem,phno]
            c=messagebox.askyesnocancel(title="print",message="do you want to print")
            if c==True or c==False:
                l="loan"
                loanprint(plist,c,l)
                loan()
                print("ok 2")
        try:
            b=int(bill)
            s=list(ws.columns)[1]
            for i in range(1,ws.max_row):
                v=str(s[i].value)
                if v==bill:
                    messagebox.showwarning(title="error",message="bill number exit")
                    break
                
            else:
                try:
                    phn=int(phno)
                    
                    if len(phno)!=10:
                        raise "billerror"
                except "billerror":
                    messagebox.showwarning(title="error",message="ph number invalid")
                    
                am=int(amount)
                w=int(weight)
                n=int(noitem)
                d=datechange(date)
                
        except:
            messagebox.showwarning(title="error",message="invalid entry")
        
        n_loan()      
                             
        
        
        

    btnEdit = Button(billing_frame, command=new_loan, text="Save", width=15, font=("Calibri", 16, "bold"),
                    fg="white", bg="#2980b9",
                    bd=0).grid(row=12, column=1, padx=10)


    btnEdit = Button(billing_frame, command=loan, text="reset", width=15, font=("Calibri", 16, "bold"),
                    fg="white", bg="#2980b9",
                    bd=0).grid(row=12, column=2, padx=10)
    

def select_tree(event):
            i=tree.item(tree.focus())
            global val
            val=i["values"]
            bill1.set(val[1])
            date1.set(val[0])
            name1.set(val[2])
            coname1.set(val[3])
            street1.set(val[4])
            address1.set(val[5])
            item1.set(val[8])
            weight1.set(val[7])
            amount1.set(val[6])
            noitem1.set(val[9])
            relese1.set(val[11])
            phvar.set(val[10])
            if val[11]==None or val[11]==NONE or val[11]=="None":
                intrel.set(tday)
            else:
                intrel.set(val[11])

def reprint():
    if val:
        c=messagebox.askyesno(title="reprint",message=f"want to reprint {val[1]}")
        if c:
            l="reprint"
            loanprint(val,c,l)
            messagebox.showinfo(title="reprint",message=f"{val[1]} reprint completed")
            detail()
    else:
        messagebox.showwarning(message="select the bill",title="x")
    
def tkdelete():
    if val:
        
        if messagebox.askyesno(title="delete",message="are you sure to delete"):
            delete(val[12])
            messagebox.showwarning(title="delete",message=f"{val[1]} deleted")
            stclear()
            detail()
            
    else:
        messagebox.showwarning(title="error",message="select bill")

  

def search():
    global search_frame
    try:
        search_frame.destroy()
        tree_frame.destroy()
        update_frame.destroy()
        tkint_frame.destroy()
    except:
        pass
    try:
        tkint_frame.destroy()
    except:
        pass
    try:
        billing_frame.destroy()
    except:
        pass
    
    stclear()
        
    search_frame=Frame(bill,bg="#535c68")
    search_frame.pack(side=TOP)
    global comb,combtxt
    comb=ttk.Combobox(search_frame,width=20,state="readonly",textvariable="stringVar()",font=("Calibri",16, "bold"),height=10)
   
    comb["values"]=["கடன் தேதி","கடன் எண்","பெயர்","த/க பெயர்","ஊர்","கடன் தொகை","எடை","பொருள்","Phone No","மீட்ட தேதி"]
    comb.grid(row=1,columnspan=2,padx=20,pady=20)
    comb.set("கடன் எண்")

    combtxt=Entry(search_frame,font=lblfont,width=20)
    combtxt.grid(row=1,column=3,pady=5)

    btnEdit = Button(search_frame, command=detail, text="search", width=15, font=("Calibri", 16, "bold"),
                 fg="white", bg="#2980b9",
                 bd=0).grid(row=1, column=4, padx=10)
    

def detail():
        global tree_frame
        
        try:
            tree_frame.destroy()
            update_frame.destroy()
        except:
            pass
        try:
            tkint_frame.destroy()  
        except:
            pass

        stclear()

        id=comb.get()
        
        c_val=combtxt.get()
        lis=srch(id,c_val)
        style=ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",highlightthicknes=0,font=('calibri',13),)
        style.configure("Treeview.Heading",highlightthicknes=0,font=('calibri',13,'bold'))
        
        tree_frame=Frame(bill,bg="#535c68")
        tree_frame.pack(side=TOP,fill="both",expand=False)
        treescroll=ttk.Scrollbar(tree_frame,orient="vertical")
        treescroll.pack(side="right",fill="y")
        
        # style=ttk.Style()
        # style.configure("mystyle.treeview",font=("Calibri", 16, "bold"),rowheight=50)
        global tree
        tree=ttk.Treeview(tree_frame,columns=(1,2,3,4,5,6,7,8,9,10,11,12),yscrollcommand=treescroll.set)
        v=["கடன் தேதி","கடன் எண்","பெயர்","த/க பெயர்","தெரு","ஊர்","கடன் தொகை","எடை","பொருள்","மொத்த பொருள்","Phone No","மீட்ட தேதி"]
        col=[1,2,3,4,5,6,7,8,9,10,11,12]
        for i in range(0,12):
            tree.heading(f"{col[i]}",text=v[i])
            tree.column(f"{col[i]}",width=5)
        tree["show"]="headings"
        tree.pack(fill="both")
        treescroll.config(command=tree.yview)
        

        up=LabelFrame(tree_frame,text="")
        up.pack(side="bottom",fill="both")
        upbutton=Button(up,text="update",command=tkupdate).grid(row=0,column=0,padx=5)
        delbutton=Button(up,text="delete",command=tkdelete).grid(row=0,column=1,padx=5)
        intbutton=Button(up,text="interest",command=tkinterest).grid(row=0,column=2,padx=5)
        pributton=Button(up,text="reprint",command=reprint).grid(row=0,column=3,padx=5)
        for i in lis:
            try:
                i[0]=datechange(i[0])
                i[11]=datechange(i[11])
            except:
                pass
            # if i==None:
            #     i="none"
            tree.insert('',END,values=i)

        tree.bind("<Double-1>", select_tree)

    


def tkupdate():
    global update_frame
    try:
        billing_frame.destroy()
    except:
        pass
    try:
        tkint_frame.destroy()
    except:
        pass
     
    

    update_frame=Frame(bill,bg="#535c68")
    update_frame.pack(side="bottom",fill=X,pady=(5,20),padx=(5,10))

    title= Label(update_frame, text="update", font=( "Calibri", 16, "bold"),bg="#535c68")
    title.grid(row=0, columnspan=2,padx=5,pady=5)
        

    lblbill=Label(update_frame,text="கடன் எண்",font=lblfont,bg="#535c68")
    lblbill.grid(row=1,column=1,pady=5,padx=10)
    txtbill=Entry(update_frame,font=lblfont,width=20,textvariable=bill1)
    txtbill.grid(row=1,column=2,pady=5,padx=10)

    lbldate=Label(update_frame,text="தேதி",font=lblfont,bg="#535c68")
    lbldate.grid(row=1,column=3,pady=5,padx=10)
    txtdate=Entry(update_frame,font=lblfont,width=20,textvariable=date1)
    txtdate.grid(row=1,column=4,pady=5,padx=10)

    lblname=Label(update_frame,text="பெயர்",font=lblfont,bg="#535c68")
    lblname.grid(row=2,column=1,pady=5,padx=10)
    txtname=Entry(update_frame,font=lblfont,width=20,textvariable=name1)
    txtname.grid(row=2,column=2,pady=5,padx=10)


    lblconame=Label(update_frame,text="த/க பெயர்",font=lblfont,bg="#535c68")
    lblconame.grid(row=2,column=3,pady=5,padx=10)
    txtconame=Entry(update_frame,font=lblfont,width=20,textvariable=coname1)
    txtconame.grid(row=2,column=4,pady=5,padx=10)

    lblstreet=Label(update_frame,text="தெரு",font=lblfont,bg="#535c68")
    lblstreet.grid(row=6,column=1,pady=5,padx=10)
    txtstreet=Entry(update_frame,font=lblfont,width=20,textvariable=street1)
    txtstreet.grid(row=6,column=2,pady=5,padx=10)

    lbladress=Label(update_frame,text="ஊர்",font=lblfont,bg="#535c68")
    lbladress.grid(row=3,column=1,pady=5,padx=10)
    txtadress=Entry(update_frame,font=lblfont,width=20,textvariable=address1)
    txtadress.grid(row=3,column=2,pady=5,padx=10)

    lblitem=Label(update_frame,text="பொருள்",font=lblfont,bg="#535c68")
    lblitem.grid(row=3,column=3,pady=5,padx=10)
    txtitem=Entry(update_frame,font=lblfont,width=20,textvariable=item1)
    txtitem.grid(row=3,column=4,pady=5,padx=10)

    lblweight=Label(update_frame,text="எடை",font=lblfont,bg="#535c68")
    lblweight.grid(row=4,column=1,pady=5,padx=10)
    txtweight=Entry(update_frame,font=lblfont,width=20,textvariable=weight1)
    txtweight.grid(row=4,column=2,pady=5,padx=10)

    lblamount=Label(update_frame,text="கடன் தொகை",font=lblfont,bg="#535c68")
    lblamount.grid(row=4,column=3,pady=5,padx=10)
    txtamount=Entry(update_frame,font=lblfont,width=20,textvariable=amount1)
    txtamount.grid(row=4,column=4,pady=5,padx=10)

    lblnoitem=Label(update_frame,text="மொத்த பொருள்",font=lblfont,bg="#535c68")
    lblnoitem.grid(row=5,column=1,pady=5,padx=10)
    txtnoitem=Entry(update_frame,font=lblfont,width=20,textvariable=noitem1)
    txtnoitem.grid(row=5,column=2,pady=5,padx=10)

    lblph=Label(update_frame,text="Phone No",font=lblfont,bg="#535c68")
    lblph.grid(row=6,column=1,pady=5,padx=10)
    txtph=Entry(update_frame,font=lblfont,width=20,textvariable=phvar)
    txtph.grid(row=6,column=2,pady=5,padx=10)

    lblrelese=Label(update_frame,text="மீட்ட தேதி",font=lblfont,bg="#535c68")
    lblrelese.grid(row=5,column=3,pady=5,padx=10)
    txtrelese=Entry(update_frame,font=lblfont,width=20,textvariable=relese1)
    txtrelese.grid(row=5,column=4,pady=5,padx=10)
   

    def updat():
        bill=txtbill.get()
        date=datechange(txtdate.get())
        name=txtname.get()
        coname=txtconame.get()
        street=txtstreet.get()
        address=txtadress.get()
        item=txtitem.get()
        weight=txtweight.get()
        amount=txtamount.get()
        noitem=txtnoitem.get()
        phno=txtph.get()
        relese=datechange(txtrelese.get())

        dlist=[date,bill,name,coname,street,address,amount,weight,item,noitem,phno,relese]
        if messagebox.askyesno(title="update",message="conform to update"):
            alter(val[12],dlist)
            messagebox.showinfo(title="update",message="update copleted")
            stclear()
            detail()
    
        



    btnEdit = Button(update_frame, command=updat, text="update", width=15, font=("Calibri", 16, "bold"),
                    fg="white", bg="#2980b9",
                    bd=0).grid(row=6, column=4, padx=10)


       
def tkinterest():
    global tkint_frame
    try:
        billing_frame.destroy()
    except:
        pass
    try:
        update_frame.destroy()
    except:
        pass
    try:
        tkint_frame.destroy()
    except:
        pass

    
   
    
    tkint_frame=Frame(bill,bg="#535c68")
    tkint_frame.pack(side="bottom",fill=X,pady=(5,10))

    title= Label(tkint_frame, text="interest", font=( "Calibri", 16, "bold"),bg="#535c68")
    title.grid(row=0, columnspan=2,padx=5,pady=5)
        

    lblbill=Label(tkint_frame,text="கடன் எண்",font=lblfont,bg="#535c68")
    lblbill.grid(row=1,column=1)
    txtbill=Entry(tkint_frame,font=lblfont,width=20,textvariable=bill1)
    txtbill.grid(row=1,column=2,pady=5)

    lbldate=Label(tkint_frame,text="தேதி",font=lblfont,bg="#535c68")
    lbldate.grid(row=1,column=3)
    txtdate=Entry(tkint_frame,font=lblfont,width=20,textvariable=date1)
    txtdate.grid(row=1,column=4,pady=5)

    lblname=Label(tkint_frame,text="பெயர்",font=lblfont,bg="#535c68")
    lblname.grid(row=2,column=1)
    txtname=Entry(tkint_frame,font=lblfont,width=20,textvariable=name1)
    txtname.grid(row=2,column=2,pady=5)


    lblconame=Label(tkint_frame,text="த/க பெயர்",font=lblfont,bg="#535c68")
    lblconame.grid(row=2,column=3)
    txtconame=Entry(tkint_frame,font=lblfont,width=20,textvariable=coname1)
    txtconame.grid(row=2,column=4,pady=5)

    lbladress=Label(tkint_frame,text="ஊர்",font=lblfont,bg="#535c68")
    lbladress.grid(row=3,column=1)
    txtadress=Entry(tkint_frame,font=lblfont,width=20,textvariable=address1)
    txtadress.grid(row=3,column=2,pady=5)

    lblitem=Label(tkint_frame,text="பொருள்",font=lblfont,bg="#535c68")
    lblitem.grid(row=3,column=3)
    txtitem=Entry(tkint_frame,font=lblfont,width=20,textvariable=item1)
    txtitem.grid(row=3,column=4,pady=5)

    lblweight=Label(tkint_frame,text="எடை",font=lblfont,bg="#535c68")
    lblweight.grid(row=4,column=1)
    txtweight=Entry(tkint_frame,font=lblfont,width=20,textvariable=weight1)
    txtweight.grid(row=4,column=2,pady=5)

    lblamount=Label(tkint_frame,text="Amount",font=lblfont,bg="#535c68")
    lblamount.grid(row=4,column=3)
    txtamount=Entry(tkint_frame,font=lblfont,width=20,textvariable=amount1)
    txtamount.grid(row=4,column=4,pady=5)

    lblnoitem=Label(tkint_frame,text="மொத்த பொருள்",font=lblfont,bg="#535c68")
    lblnoitem.grid(row=5,column=1)
    txtnoitem=Entry(tkint_frame,font=lblfont,width=20,textvariable=noitem1)
    txtnoitem.grid(row=5,column=2,pady=5)

    lbltoday=Label(tkint_frame,text="மீட்பு தேதி",font=lblfont,bg="#535c68")
    lbltoday.grid(row=5,column=3)
    txttoday=Entry(tkint_frame,font=lblfont,width=20,textvariable=intrel)
    txttoday.grid(row=5,column=4,pady=5)

    intpay=StringVar()
    totpay=StringVar()
        
    
    

    def intcal():
        bill=txtbill.get()
        date=datechange(txtdate.get())
        name=txtname.get()
        coname=txtconame.get()
        address=txtadress.get()
        item=txtitem.get()
        weight=txtweight.get()
        amount=txtamount.get()
        noitem=txtnoitem.get()
        redate=datechange(txttoday.get())
        if redate==None:
            messagebox.showwarning(title="release date",message="enter release date")

        dlist=[date,bill,name,coname,address,amount,weight,item,noitem]
        interestamt=interest(amount,date,redate)
        
        
        lblinterest=Label(tkint_frame,text="interest",font=lblfont,bg="#535c68")
        lblinterest.grid(row=1,column=6)
        txtinterest=Entry(tkint_frame,font=lblfont,width=20,textvariable=intpay)
        txtinterest.grid(row=1,column=7,pady=5)

        lbltotal=Label(tkint_frame,text="total",font=lblfont,bg="#535c68")
        lbltotal.grid(row=2,column=6)
        txttotal=Label(tkint_frame,font=lblfont,width=20)
        txttotal.grid(row=2,column=7,pady=5)
        totl=int(interestamt)+int(amount)
        intpay.set(interestamt)
        txttotal['text']=totl
        def tt():
            global totl
            totl=int(txtinterest.get())+int(amount)
            txttotal['text']=totl
            


        btnint=Button(tkint_frame,command=tt,text="total",width=5).grid(row=1,column=8)
        
        
        # paylist=[name,date,bill,amount,redate,int(txtinterest.get()),totl,val[12]]
        
        def payint():
            paylist=[name,date,bill,amount,redate,int(txtinterest.get()),txttotal['text'],val[12]]
        
            c=messagebox.askyesno(title="print",message="do  you want to print")
            interestprint(paylist,c)
            search()
            

        btnEdit = Button(tkint_frame, command=payint, text="pay", width=15, font=("Calibri", 16, "bold"),
                    fg="white", bg="#2980b9",
                    bd=0).grid(row=3, column=7, padx=10)
        
        


    btnEdit = Button(tkint_frame, command=intcal, text="interest", width=15, font=("Calibri", 16, "bold"),
                    fg="white", bg="#2980b9",
                    bd=0).grid(row=6, column=4, padx=10)





btnEdit = Button(head_frame, command=loan, text="loan", width=15, font=("Calibri", 16, "bold"),height=-5,
                 fg="white", bg="#2980b9",
                 bd=0).grid(row=1, column=1,pady=1,padx=10)


btnEdit = Button(head_frame, command=search, text="search", width=15, font=("Calibri", 16, "bold"),height=-5,
                 fg="white", bg="#2980b9",
                 bd=0).grid(row=1, column=2,pady=1,padx=10)


bill.mainloop()